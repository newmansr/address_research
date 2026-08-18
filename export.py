"""
export.py — write per-unit results to CSV or XLSX (one row per unit).

Format is chosen by the output path's extension (.csv or .xlsx). CSV uses only the
stdlib; XLSX uses openpyxl (with a frozen header row + auto-ish column widths).
"""

from __future__ import annotations

import csv
import os

COLUMNS = [
    ("unit", "Unit"),
    ("most_likely", "Most Likely Resident"),
    ("confidence", "Confidence"),
    ("age", "Age"),
    ("since", "Resident Since"),
    ("phones", "Phones"),
    ("sources", "Corroborating Sources"),
    ("score", "Score"),
    ("other_candidates", "Other Candidates"),
    ("possible", "Possible (unconfirmed)"),
    ("former", "Former Residents"),
    ("owner", "Owner of Record"),
    ("building_only", "Building-only (unit unknown)"),
]


def build_row(unit: str, current, possible, former, building_only: int, owner: str = "") -> dict:
    """Flatten one unit's scored result into a single spreadsheet row."""
    top = current[0] if current else None
    others = current[1:] if current else []
    return {
        "unit": unit or "(building)",
        "most_likely": top.name if top else "",
        "confidence": top.confidence if top else "none",
        "age": top.age if (top and top.age) else "",
        "since": top.since if (top and top.since) else "",
        "phones": ", ".join(top.phones) if top else "",
        "sources": ", ".join(top.sources) if top else "",
        "score": top.score if top else "",
        "other_candidates": "; ".join(f"{c.name} [{c.confidence}]" for c in others),
        # Surface the unconfirmed leads only when there's no confirmed current resident.
        "possible": "" if current else "; ".join(f"{c.name} [{', '.join(c.sources)}]"
                                                  for c in possible),
        "former": ", ".join(c.name for c in former),
        "owner": owner,
        "building_only": building_only,
    }


def write_results(path: str, rows: list[dict]) -> str:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        _write_xlsx(path, rows)
    else:
        if ext != ".csv":
            path += ".csv"
        _write_csv(path, rows)
    return path


def _write_csv(path: str, rows: list[dict]) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow([label for _, label in COLUMNS])
        for r in rows:
            w.writerow([r.get(key, "") for key, _ in COLUMNS])


def _write_xlsx(path: str, rows: list[dict]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Residents"
    ws.append([label for _, label in COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for r in rows:
        ws.append([r.get(key, "") for key, _ in COLUMNS])

    # Reasonable column widths.
    widths = {"Most Likely Resident": 26, "Phones": 34, "Corroborating Sources": 26,
              "Other Candidates": 40, "Former Residents": 30, "Owner of Record": 44,
              "Building-only (unit unknown)": 14}
    for i, (_, label) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = widths.get(label, 12)

    wb.save(path)
